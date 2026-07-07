"""
Transforms the monthly wide-format Basti raw files (one row per school, one column
per competency) into a single long-format Raw_Scores table (Month, Block, School,
competencies...) plus a Competency_Meta reference table.

This is the Google-Sheets-ready data model: adding a new month means appending rows
to Raw_Scores with a new Month value -- no new sheets, no new columns, no formula
rewrites required.

Usage: py build_long_format.py
Reads the month files listed in MONTHS below and writes:
  - Raw_Scores.csv
  - Competency_Meta.csv
  - Basti_Sheets_Import.xlsx  (both tables as tabs, ready for Google Sheets import)
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

RAW_COLS = ['dist_name','BlockTownName','Arrtype','DISE_SchoolCode','School_Name','Total','NL_SchoolCode',
            'H104.2','H106.1','H108.1','M 101.1 (A)','M 101.1 (B)','M101.2','M102.1','M102.3',
            'H202.2','H106.2','H108.3','M201.2','M201.3','M103.1/M103.2','M205.1','M207.1',
            'H208','H211','M301.1','M301.2','M301.4','M302.1','H301']

COMPETENCIES = [
    ('H104.2', 'Oral Expression', 'G1', 'Literacy'),
    ('H106.1', 'Word Reading', 'G1', 'Literacy'),
    ('H108.1', 'Letter Writing', 'G1', 'Literacy'),
    ('M 101.1 (A)', 'Spatial Concepts', 'G1', 'Numeracy'),
    ('M 101.1 (B)', 'Spatial Concepts', 'G1', 'Numeracy'),
    ('M101.2', 'Pre-Number Concepts', 'G1', 'Numeracy'),
    ('M102.1', 'Number Recognition (1-99)', 'G1', 'Numeracy'),
    ('M102.3', 'Counting 1-20', 'G1', 'Numeracy'),
    ('H202.2', 'Oral speaking', 'G2', 'Literacy'),
    ('H106.2', 'Sentence Reading', 'G2', 'Literacy'),
    ('H108.3', 'Sentence Writing', 'G2', 'Literacy'),
    ('M201.2', 'Place Value 1-99', 'G2', 'Numeracy'),
    ('M201.3', 'Number Comparison', 'G2', 'Numeracy'),
    ('M103.1/M103.2', 'Addition & Subtraction', 'G2', 'Numeracy'),
    ('M205.1', 'Shapes', 'G2', 'Numeracy'),
    ('M207.1', 'Patterns', 'G2', 'Numeracy'),
    ('H208', 'Reading Comprehension', 'G3', 'Literacy'),
    ('H211', 'Sentence Writing (Descriptive)', 'G3', 'Literacy'),
    ('M301.1', 'Number Recognition', 'G3', 'Numeracy'),
    ('M301.2', 'Place Value 1-999', 'G3', 'Numeracy'),
    ('M301.4', 'Number Comparison', 'G3', 'Numeracy'),
    ('M302.1', 'Operations with Zero', 'G3', 'Numeracy'),
    ('H301', 'Listening Skills', 'G3', 'Literacy'),
]

MONTHS = [
    ('2026-04', 'April', r'C:\Users\adich\Downloads\April Basti.xlsx', 'Basti'),
    ('2026-05', 'May', r'C:\Users\adich\Downloads\May Basti.xlsx', 'Basti(1)'),
    ('2026-06', 'June', r'C:\Users\adich\Downloads\June Basti.xlsx', 'LOBasti'),
]

frames = []
for month_key, month_label, path, sheet in MONTHS:
    df = pd.read_excel(path, sheet_name=sheet)
    df = df[RAW_COLS].copy()
    df.insert(0, 'MonthLabel', month_label)
    df.insert(0, 'Month', month_key)
    frames.append(df)

long_df = pd.concat(frames, ignore_index=True)
long_df = long_df.drop(columns=['dist_name', 'NL_SchoolCode'])
long_df = long_df.rename(columns={'BlockTownName': 'Block', 'DISE_SchoolCode': 'SchoolCode', 'School_Name': 'School'})

print('Long format shape:', long_df.shape)
print(long_df['Month'].value_counts())
print(long_df.head(3).to_string())

out_dir = r'C:\Users\adich\Desktop\basti-nipun-dashboard\data-prep'
long_df.to_csv(f'{out_dir}\\Raw_Scores.csv', index=False)

meta_df = pd.DataFrame(COMPETENCIES, columns=['Code', 'Desc', 'Grade', 'Subject'])
meta_df.to_csv(f'{out_dir}\\Competency_Meta.csv', index=False)

# Build a combined xlsx for easy Google Sheets import (File > Import > Replace spreadsheet, or per-tab import)
wb = Workbook()
ws1 = wb.active
ws1.title = 'Raw_Scores'
for r in dataframe_to_rows(long_df, index=False, header=True):
    ws1.append(r)
for cell in ws1[1]:
    cell.font = Font(bold=True, color='FFFFFFFF')
    cell.fill = PatternFill('solid', fgColor='FF1F4E79')
ws1.freeze_panes = 'A2'

ws2 = wb.create_sheet('Competency_Meta')
for r in dataframe_to_rows(meta_df, index=False, header=True):
    ws2.append(r)
for cell in ws2[1]:
    cell.font = Font(bold=True, color='FFFFFFFF')
    cell.fill = PatternFill('solid', fgColor='FF1F4E79')
ws2.freeze_panes = 'A2'

wb.save(f'{out_dir}\\Basti_Sheets_Import.xlsx')
print('Saved CSVs and import xlsx to', out_dir)
